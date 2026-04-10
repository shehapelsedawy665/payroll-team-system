"""
Tax Portal Export — نموذج القطاع الخاص 2026
Generates exact-format Excel file from MongoDB payroll data.
Usage: python generate_tax_export.py <month YYYY-MM> <output_path>
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ─── Column definitions: (field_code, arabic_name, mongo_path_or_constant) ───
COLUMNS = [
    ("EI005",  "مسلسل",                                    "seq"),
    ("EI010",  "كود الموظف",                               "employee._id"),
    ("EI015",  "اسم الموظف",                               "employee.name"),
    ("EI020",  "الجنسية",                                  "employee.nationality|مصري"),
    ("EI025",  "الرقم القومي",                             "employee.nationalId"),
    ("EI026",  "رقم جواز السفر",                           "employee.passportNo|"),
    ("EI027",  "رقم جواز السفر الجديد",                   "employee.passportNoNew|"),
    ("EI030",  "رقم التليفون",                             "employee.phone|"),
    ("EI035",  "حالة تصريح العمل لغير المصريين",          "const|"),
    ("EI040",  "رقم تصريح العمل",                          "const|"),
    ("EI045",  "الوظيفة",                                  "employee.position|"),
    ("EI055",  "اسم الجهة/الفرع",                          "company.name"),
    ("EI060",  "المعاملة الضريبية",                        "const|1"),
    ("EI065",  "رقم التسجيل الضريبي لجهة العمل الأصلية", "company.taxRegNo|"),
    ("EI130",  "مدة العمل",                                "payload.days"),
    ("EI070",  "الحالة التأمينية",                         "const|1"),
    ("EI075",  "الرقم التأمينى",                           "employee.insuranceNo|"),
    ("EI080",  "تاريخ الالتحاق بالتأمينات",               "employee.hiringDate"),
    ("EI085",  "قسط مدة سابقة",                            "const|0"),
    ("EI090",  "تاريخ نهاية الخدمة",                       "employee.resignationDate|"),
    ("EI095",  "تاريخ انتهاء الاشتراك من التأمينات",      "employee.resignationDate|"),
    ("EI100",  "الأجر الشامل",                             "payload.gross"),
    ("EI105",  "بدلات غير خاضعة تأمينيأ",                 "const|0"),
    ("EI110",  "الأجر التأميني",                           "payload.insBase"),
    ("EI115",  "حالة التأمين الصحي الشامل",               "const|1"),
    ("EI120",  "عدد الزوجات الغير عاملات",                 "const|0"),
    ("EI125",  "عدد المعالين",                              "const|0"),
    ("DTE160", "المرتب الأساسي",                           "payload.proratedBasic"),
    ("DTE170", "مكافات وحوافز/أجر إضافي/منح",             "payload.additions_bonus|0"),
    ("DTE180", "علاوات خاصة معفاة",                        "payload.additions_exempt|0"),
    ("DTE175", "علاوات خاصة خاضعة",                       "payload.additions_taxable|0"),
    ("DTE190", "عمولات",                                   "const|0"),
    ("DTE230", "نصيب العامل في الأرباح",                   "const|0"),
    ("DTE235", "مقابل الخدمة",                             "const|0"),
    ("DTE260", "البقشيش",                                  "const|0"),
    ("DTE240", "مرتبات رؤساء مجلس الادارة",               "const|0"),
    ("DTE245", "المقابل النقدى لرصيد الاجازات",           "const|0"),
    ("DTE250", "مكافأة نهاية الخدمة الخاضعة",             "const|0"),
    ("DTE255", "مبالغ منصرفة بقوانين خاصة",               "const|0"),
    ("DTE265", "إضافات وبدلات اخرى خاضعة",               "payload.proratedTrans"),
    ("DTE270", "ما تحملته المنشاة من ضريبة مرتبات",       "const|0"),
    ("DTE275", "ما تحملته المنشأه من حصة العامل في التأمينات", "const|0"),
    ("DTE280", "مبالغ خاضعة ربع سنوية",                   "const|0"),
    ("DTE285", "مبالغ خاضعة نصف سنوية",                   "const|0"),
    ("DTE290", "مبالغ خاضعة سنوية",                       "const|0"),
    ("DTE200", "مزايا: السيارات",                          "const|0"),
    ("DTE205", "مزايا: الهواتف المحمولة",                  "const|0"),
    ("DTE210", "مزايا: قروض وسلف",                        "const|0"),
    ("DTE215", "مزايا: التأمين على الحياة",               "const|0"),
    ("DTE220", "مزايا: اسهم الشركة",                      "const|0"),
    ("DTE225", "مزايا أخرى",                               "const|0"),
    ("DTE295", "اجمالى الاستحقاقات",                      "payload.gross"),
    ("DAE405", "حصة العامل في التأمينات",                  "payload.insuranceEmployee"),
    ("DAE410", "حصة العامل في التأمين الصحي",             "const|0"),
    ("DAE415", "مبالغ معفاة بقوانين خاصة",                "const|0"),
    ("DAE420", "علاوات خاصة معفاة",                        "const|0"),
    ("DAE425", "العلاوة الاجتماعية",                       "const|0"),
    ("DAE430", "الاعفاء الشخصى",                           "payload.personalExemptionUsed|0"),
    ("DAE435", "اقساط مدة سابقة",                          "const|0"),
    ("DAE440", "نصيب العامل في الأرباح",                   "const|0"),
    ("DAE450", "اشتراكات صناديق التامين (ق54/75)",        "const|0"),
    ("DAE451", "اشتراكات صناديق التامين (ق155/2024)",     "const|0"),
    ("DAE455", "أقساط التأمين على الحياة",                 "const|0"),
    ("DAE460", "أقساط التأمين الصحي",                      "const|0"),
    ("DAE465", "أقساط تأمين لاستحقاق معاش",               "const|0"),
    ("DAE470", "إجمالي اشتراكات صناديق التأمين",          "const|0"),
    ("DAE475", "اجمالى الاستقطاعات",                      "payload.totalAllDeductions"),
    ("TC505",  "صافى الدخل (وعاء الفتره)",                "payload.currentTaxable"),
    ("TC510",  "الوعاء السنوى",                            "payload.annualProjected"),
    ("TC515",  "الضريبة المستحقة (عمالة أصلية)",          "payload.monthlyTax"),
    ("TC520",  "الضريبة المستحقة (نموذج 3)",               "const|0"),
    ("TC525",  "الضريبة المستحقة (نموذج 2)",               "const|0"),
    ("TC530",  "اجمالى الضريبة المستحقة",                  "payload.monthlyTax"),
    ("END705", "الضريبة المحتسبة عن الفترة",              "payload.monthlyTax"),
    ("TC535",  "ضريبة فترات سابقة (معاملة 1,4,5,6,7)",   "payload.prevTaxes"),
    ("TC540",  "ضريبة الفترة (معاملة 1,4,5,6,7)",         "payload.monthlyTax"),
    ("TC545",  "ضريبة فترات سابقة (معاملة 2)",            "const|0"),
    ("TC550",  "ضريبة الفترة (معاملة 2)",                 "const|0"),
    ("TC555",  "ضريبة فترات سابقة (معاملة 3)",            "const|0"),
    ("TC560",  "ضريبة الفترة (معاملة 3)",                 "const|0"),
    ("TC565",  "صافي الأجر النهائي",                       "payload.net"),
    ("NAD610", "صندوق الشهداء",                            "payload.martyrs"),
    ("CLD825", "دعم ذوي الهمم",                            "const|0"),
    ("NAD615", "اضافات: السلفة/قروض",                     "const|0"),
    ("NAD620", "اضافات: مكافأة نهاية الخدمة غير خاضعة",  "const|0"),
    ("NAD625", "اضافات: رصيد الأجازات غير خاضعة",        "const|0"),
    ("NAD630", "اضافات أخرى",                              "const|0"),
    ("NAD635", "استقطاعات: نفقة",                         "const|0"),
    ("NAD640", "استقطاعات: قسط السلفة/القرض",             "const|0"),
    ("NAD645", "استقطاعات: اشتراكات نقابات",              "const|0"),
    ("NAD650", "استقطاعات: جزاءات",                       "const|0"),
    ("NAD655", "استقطاعات: قسط بوليصة التأمين",           "const|0"),
    ("NAD660", "استقطاعات أخرى",                          "const|0"),
    ("CLD805", "حصة الشركة في التأمينات",                  "company.employerInsRate|0"),
    ("CLD810", "حصة الشركة في التأمين الصحي",             "const|0"),
    ("CLD815", "المساهمة في صندوق الشهداء",               "payload.martyrs"),
    ("CLD820", "الدمغات",                                  "const|0"),
    ("NAD665", "المبالغ المحولة فعلياً",                   "payload.net"),
]

def get_val(record, path, seq=None):
    """Resolve a value from path notation."""
    if path == "seq":
        return seq
    if path.startswith("const|"):
        v = path.split("|", 1)[1]
        try: return float(v) if v else ""
        except: return v
    
    parts = path.split("|")
    key = parts[0]
    default = parts[1] if len(parts) > 1 else ""
    
    segments = key.split(".")
    val = record
    for s in segments:
        if isinstance(val, dict):
            val = val.get(s, default)
        else:
            return default
    
    if val is None or val == "":
        return default if default != "" else 0
    return val


def validate_records(records):
    """Validate records before export. Returns list of error strings."""
    errors = []
    required_per_record = ["employee.nationalId", "employee.name", "payload.gross", "payload.net"]
    for i, r in enumerate(records):
        name = get_val(r, "employee.name", i+1)
        if not get_val(r, "employee.nationalId"):
            errors.append(f"صف {i+1} ({name}): الرقم القومي مطلوب")
        gross = get_val(r, "payload.gross")
        net   = get_val(r, "payload.net")
        try:
            if float(gross) < float(net):
                errors.append(f"صف {i+1} ({name}): الشامل ({gross}) أقل من الصافي ({net})")
        except: pass
    return errors


def build_excel(records, month, output_path, company_name=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج القطاع الخاص"
    ws.sheet_view.rightToLeft = True

    # ── Header row 1: Arabic names ──
    header_fill = PatternFill("solid", start_color="1E3A5F")
    code_fill   = PatternFill("solid", start_color="2D6A9F")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
    code_font   = Font(name="Arial", bold=True, color="E8F4FD", size=7)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    for col_idx, (code, arabic, _) in enumerate(COLUMNS, start=1):
        c1 = ws.cell(row=1, column=col_idx, value=arabic)
        c1.fill = header_fill; c1.font = header_font
        c1.alignment = center_al; c1.border = thin_border

        c2 = ws.cell(row=2, column=col_idx, value=code)
        c2.fill = code_fill; c2.font = code_font
        c2.alignment = center_al; c2.border = thin_border

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 18

    # ── Data rows ──
    data_fill_even = PatternFill("solid", start_color="F0F7FF")
    data_font      = Font(name="Arial", size=9)
    num_format     = '#,##0.00'
    int_format     = '#,##0'

    numeric_cols = set()
    for i, (code, _, path) in enumerate(COLUMNS):
        if not (path.startswith("const|") and not path.split("|")[1].replace(".","").isdigit()):
            try:
                float(path.split("|")[-1])
                numeric_cols.add(i+1)
            except: pass

    for row_idx, record in enumerate(records, start=3):
        fill = data_fill_even if row_idx % 2 == 0 else None
        for col_idx, (code, arabic, path) in enumerate(COLUMNS, start=1):
            val = get_val(record, path, seq=row_idx - 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill: cell.fill = fill
            
            # Number formatting for financial columns
            if isinstance(val, (int, float)) and col_idx > 13:
                cell.number_format = num_format

    # ── Column widths ──
    ws.column_dimensions['A'].width = 6   # seq
    ws.column_dimensions['B'].width = 10  # employee code
    ws.column_dimensions['C'].width = 22  # name
    ws.column_dimensions['E'].width = 16  # national ID
    for i in range(4, len(COLUMNS)+1):
        col_letter = get_column_letter(i)
        if col_letter not in ('A','B','C','E'):
            ws.column_dimensions[col_letter].width = 14

    # ── Freeze panes ──
    ws.freeze_panes = "D3"

    # ── Summary row ──
    sum_row = len(records) + 3
    sum_fill = PatternFill("solid", start_color="FFF3CD")
    sum_font = Font(name="Arial", bold=True, size=9, color="7B4700")
    ws.cell(row=sum_row, column=1, value="الإجمالي").font = sum_font
    ws.cell(row=sum_row, column=1).fill = sum_fill
    ws.cell(row=sum_row, column=1).alignment = center_al

    # Sum financial columns
    financial_col_indices = [21,22,23,27,28,51,52,66,67,68,69,72,73,80,81,93,97]
    for cidx in financial_col_indices:
        if cidx <= len(COLUMNS):
            col_letter = get_column_letter(cidx)
            formula = f"=SUM({col_letter}3:{col_letter}{sum_row-1})"
            c = ws.cell(row=sum_row, column=cidx, value=formula)
            c.font = sum_font; c.fill = sum_fill
            c.number_format = num_format
            c.alignment = center_al

    # ── Title sheet info ──
    ws.cell(row=sum_row+2, column=1, value=f"الشركة: {company_name}").font = Font(bold=True, size=10)
    ws.cell(row=sum_row+3, column=1, value=f"الفترة: {month}").font = Font(bold=True, size=10)
    ws.cell(row=sum_row+4, column=1, value=f"عدد الموظفين: {len(records)}").font = Font(bold=True, size=10)

    wb.save(output_path)
    print(f"✅ Tax portal file saved: {output_path} ({len(records)} employees)")


# ── DEMO: Generate sample file ──
if __name__ == "__main__":
    SAMPLE_RECORDS = [
        {
            "employee": {
                "_id": "EMP001",
                "name": "أحمد محمد علي",
                "nationalId": "29901011234567",
                "phone": "01001234567",
                "position": "محاسب",
                "hiringDate": "2020-01-01",
                "resignationDate": "",
                "insuranceNo": "12345678",
                "nationality": "مصري"
            },
            "company": {
                "name": "شركة تطوير الأعمال",
                "taxRegNo": "123456789",
                "employerInsRate": 1837
            },
            "payload": {
                "days": 30,
                "proratedBasic": 8000,
                "proratedTrans": 1500,
                "gross": 9500,
                "insBase": 9500,
                "insuranceEmployee": 1045,
                "currentTaxable": 5650,
                "annualProjected": 67800,
                "monthlyTax": 312.50,
                "prevTaxes": 0,
                "martyrs": 4.75,
                "totalAllDeductions": 1362.25,
                "net": 8137.75,
                "additions_bonus": 0,
                "additions_exempt": 0,
                "additions_taxable": 0
            }
        },
        {
            "employee": {
                "_id": "EMP002",
                "name": "سارة إبراهيم حسن",
                "nationalId": "29503152345678",
                "phone": "01112345678",
                "position": "مهندس برمجيات",
                "hiringDate": "2021-03-15",
                "resignationDate": "",
                "insuranceNo": "23456789",
                "nationality": "مصري"
            },
            "company": {
                "name": "شركة تطوير الأعمال",
                "taxRegNo": "123456789",
                "employerInsRate": 2695
            },
            "payload": {
                "days": 30,
                "proratedBasic": 12000,
                "proratedTrans": 2000,
                "gross": 14000,
                "insBase": 14000,
                "insuranceEmployee": 1540,
                "currentTaxable": 9655,
                "annualProjected": 115860,
                "monthlyTax": 1135.75,
                "prevTaxes": 0,
                "martyrs": 7,
                "totalAllDeductions": 2682.75,
                "net": 11317.25,
                "additions_bonus": 0,
                "additions_exempt": 0,
                "additions_taxable": 0
            }
        }
    ]

    out = sys.argv[1] if len(sys.argv) > 1 else "/home/claude/erp-upgrade/tax_portal_sample.xlsx"
    month = sys.argv[2] if len(sys.argv) > 2 else "2026-03"
    
    errors = validate_records(SAMPLE_RECORDS)
    if errors:
        print("❌ Validation errors:")
        for e in errors: print(" -", e)
        sys.exit(1)
    
    build_excel(SAMPLE_RECORDS, month, out, "شركة تطوير الأعمال")


def build_from_json(json_path, output_path):
    """Called by Node.js with pre-built records array."""
    with open(json_path) as f:
        data = json.load(f)
    
    # data.records is already an array of flat row arrays
    raw_records = data.get("records", [])
    month       = data.get("month", "")
    company_nm  = data.get("company", "")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج القطاع الخاص"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", start_color="1E3A5F")
    code_fill   = PatternFill("solid", start_color="2D6A9F")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
    code_font   = Font(name="Arial", bold=True, color="E8F4FD", size=7)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin",color="CCCCCC"), right=Side(style="thin",color="CCCCCC"),
                  top=Side(style="thin",color="CCCCCC"), bottom=Side(style="thin",color="CCCCCC"))

    # Header rows
    for col_idx, (code, arabic, _) in enumerate(COLUMNS, start=1):
        c1 = ws.cell(row=1, column=col_idx, value=arabic)
        c1.fill = header_fill; c1.font = header_font; c1.alignment = center_al; c1.border = thin
        c2 = ws.cell(row=2, column=col_idx, value=code)
        c2.fill = code_fill;  c2.font = code_font;  c2.alignment = center_al; c2.border = thin

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 18

    data_font = Font(name="Arial", size=9)
    even_fill = PatternFill("solid", start_color="F0F7FF")
    num_fmt   = '#,##0.00'

    for row_idx, row_vals in enumerate(raw_records, start=3):
        fill = even_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font; cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill: cell.fill = fill
            if isinstance(val, float) and col_idx > 13:
                cell.number_format = num_fmt

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['E'].width = 16
    for i in range(4, len(COLUMNS)+1):
        ltr = get_column_letter(i)
        if ltr not in ('A','B','C','E'):
            ws.column_dimensions[ltr].width = 14

    ws.freeze_panes = "D3"

    # Summary row
    sr = len(raw_records) + 3
    sf = PatternFill("solid", start_color="FFF3CD")
    sfont = Font(name="Arial", bold=True, size=9, color="7B4700")
    ws.cell(row=sr, column=1, value="الإجمالي").font = sfont
    ws.cell(row=sr, column=1).fill = sf
    for cidx in [21,22,23,27,28,51,52,66,67,68,69,72,73,80,81,93,97]:
        if cidx <= len(COLUMNS):
            cl = get_column_letter(cidx)
            c = ws.cell(row=sr, column=cidx, value=f"=SUM({cl}3:{cl}{sr-1})")
            c.font = sfont; c.fill = sf; c.number_format = num_fmt
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=sr+2, column=1, value=f"الشركة: {company_nm}").font = Font(bold=True,size=10)
    ws.cell(row=sr+3, column=1, value=f"الفترة: {month}").font = Font(bold=True,size=10)
    ws.cell(row=sr+4, column=1, value=f"عدد الموظفين: {len(raw_records)}").font = Font(bold=True,size=10)

    wb.save(output_path)
    print(f"✅ Saved: {output_path} ({len(raw_records)} rows)")


# ── Entry point from Node.js ──
# Called as: python3 generate_tax_export.py <output.xlsx> <month> <data.json>
if len(sys.argv) == 4:
    build_from_json(sys.argv[3], sys.argv[1])
